"""
export_utils.py
---------------
Export utilities for session attendance reports.
Supports students-only, staff-only, and both sessions.
- Excel: separate sheets per type
- PDF:   separate sections per type
"""

from collections import OrderedDict
from datetime import datetime
from db.students_db import YEAR_LEVEL_LABELS
from database import (SessionLocal, Session as EventSession,
                      SessionPeriod, Attendance, Student,
                      StaffAttendance, Staff,
                      AcademicPeriod, AcademicYear, AcademicTerm)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable, KeepTogether
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas as rl_canvas


# ---------------------------------------------------------------------------
# Export data fetcher
# ---------------------------------------------------------------------------

def fetch_session_report(session_id: int) -> dict:
    db = SessionLocal()
    try:
        session = db.query(EventSession).filter_by(id=session_id).first()
        if not session:
            return {}

        atype = getattr(session, "attendee_type", "students") or "students"

        periods = (db.query(SessionPeriod)
                   .filter_by(session_id=session_id)
                   .order_by(SessionPeriod.sort_order)
                   .all())
        period_map = {p.id: p for p in periods}

        # ── Student records ───────────────────────────────────────────
        student_rows = []
        student_present = student_late = 0

        if atype in ("students", "both"):
            records = (
                db.query(Attendance, Student)
                .join(Student, Attendance.student_id == Student.student_id)
                .filter(Attendance.session_id == session_id)
                .order_by(Attendance.time_in)
                .all()
            )
            for att, stu in records:
                period_obj = period_map.get(att.period_id)
                student_rows.append({
                    "entity_id":  str(stu.student_id),
                    "name":       f"{stu.first_name} {stu.last_name}",
                    "program":    stu.program.name if stu.program else "—",
                    "code":       stu.program.code if stu.program else "—",
                    "yearlevel":  YEAR_LEVEL_LABELS.get(stu.year_level, "—"),
                    "col3":       stu.program.code if stu.program else "—",
                    "col4":       YEAR_LEVEL_LABELS.get(stu.year_level, "—"),
                    "status":     att.status,
                    "period_id":  att.period_id,
                    "period":     period_obj.name if period_obj else "—",
                    "time_in":    att.time_in.strftime("%I:%M:%S %p")  if att.time_in  else "—",
                    "time_out":   att.time_out.strftime("%I:%M:%S %p") if att.time_out else "—",
                    "department": stu.program.department.name if stu.program and stu.program.department else "—",
                    "dept_code":  stu.program.department.code if stu.program and stu.program.department else "—",
                    "type":       "student",
                })
                if att.status == "present":
                    student_present += 1
                elif att.status == "late":
                    student_late += 1
                    student_present += 1

        # ── Staff records ─────────────────────────────────────────────
        staff_rows = []
        staff_present = staff_late = 0

        if atype in ("staff", "both"):
            s_records = (
                db.query(StaffAttendance, Staff)
                .join(Staff, StaffAttendance.staff_id == Staff.staff_id)
                .filter(StaffAttendance.session_id == session_id)
                .order_by(StaffAttendance.time_in)
                .all()
            )
            for att, st in s_records:
                period_obj = period_map.get(att.period_id)
                staff_rows.append({
                    "entity_id":  str(st.staff_id),
                    "name":       f"{st.first_name} {st.last_name}",
                    "col3":       st.department,
                    "col4":       st.role,
                    "status":     att.status,
                    "period_id":  att.period_id,
                    "period":     period_obj.name if period_obj else "—",
                    "time_in":    att.time_in.strftime("%I:%M:%S %p")  if att.time_in  else "—",
                    "time_out":   att.time_out.strftime("%I:%M:%S %p") if att.time_out else "—",
                    "department": st.department,
                    "role":       st.role,
                    "type":       "staff",
                })
                if att.status == "present":
                    staff_present += 1
                elif att.status == "late":
                    staff_late += 1
                    staff_present += 1

        # ── Summary ───────────────────────────────────────────────────
        student_est = session.student_estimated or 0
        staff_est = session.staff_estimated or 0
        est = student_est + staff_est
        total_present = student_present + staff_present
        absent = max(0, est - total_present) if est else 0
        rate   = min(round(total_present / est * 100, 1), 100.0) if est else None

        summary = {
            "present":         total_present,
            "late":            student_late + staff_late,
            "absent":          absent,
            "estimated":       est,
            "rate":            rate,
            "student_present": student_present,
            "student_late":    student_late,
            "staff_present":   staff_present,
            "staff_late":      staff_late,
        }

        # ── Academic period string ─────────────────────────────────────
        academic_period_str = "—"
        if session.academic_period_id:
            ap = db.query(AcademicPeriod).filter_by(id=session.academic_period_id).first()
            if ap:
                term = db.query(AcademicTerm).filter_by(id=ap.term_id).first()
                year = db.query(AcademicYear).filter_by(id=ap.academic_year_id).first()
                if term and year:
                    academic_period_str = f"{term.name} • A.Y. {year.year_start}–{year.year_end}"

        return {
            "session":          session,
            "attendee_type":    atype,
            "student_rows":     student_rows,
            "staff_rows":       staff_rows,
            "rows":             student_rows + staff_rows,  # legacy compat
            "summary":          summary,
            "period_map":       period_map,
            "periods":          periods,
            "any_late_enabled": any(p.late_enabled for p in periods),
            "academic_period":  academic_period_str,
        }

    finally:
        db.close()


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

XL_HEADER_BG  = "1e3a5f"
XL_HEADER_FG  = "ffffff"
XL_ACCENT     = "2563eb"
XL_STUDENT    = "0369a1"   # blue — students sheet
XL_STAFF      = "7c3aed"   # purple — staff sheet
XL_SUCCESS    = "16a34a"
XL_WARNING    = "d97706"
XL_ERROR      = "dc2626"
XL_ROW_ALT    = "f1f5f9"
XL_ROW_NORMAL = "ffffff"
XL_TEXT       = "1e293b"
XL_MUTED      = "64748b"
XL_SECTION_BG = "e8f0fe"

_XL_STATUS_COLORS = {"present": XL_SUCCESS, "late": XL_WARNING, "absent": XL_ERROR}


def _xl_font(bold=False, color=XL_TEXT, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _xl_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _xl_border():
    s = Side(style="thin", color="cbd5e1")
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_bottom_border():
    b = Side(style="medium", color="2563eb")
    t = Side(style="thin", color="cbd5e1")
    return Border(left=t, right=t, top=t, bottom=b)

def _xl_center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _xl_left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _xl_autosize(ws, min_width=8, max_width=60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    factor = 1.2 if (cell.font and cell.font.bold) else 1.0
                    max_len = max(max_len, int(len(str(cell.value)) * factor))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _xl_write_header_block(ws, session, summary, num_cols: int,
                           entity_type: str = "students",
                           any_late_enabled: bool = True,
                           academic_period: str = "—") -> int:
    last_col = get_column_letter(num_cols)
    accent   = XL_STUDENT if entity_type == "students" else XL_STAFF
    label    = "STUDENTS" if entity_type == "students" else "STAFF"

    # Row 1 — Title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = f"ATTENDANCE REPORT — {label}"
    c.font      = _xl_font(bold=True, size=14, color=accent)
    c.fill      = _xl_fill(XL_HEADER_BG)
    c.alignment = _xl_center()
    ws.row_dimensions[1].height = 28

    # Row 2 — Session info
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"Session: {session.name} | Date: {session.date} | {academic_period}"
    c.font      = _xl_font(size=10, color=XL_MUTED)
    c.fill      = _xl_fill(XL_HEADER_BG)
    c.alignment = _xl_center()
    ws.row_dimensions[2].height = 18

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].fill = _xl_fill(XL_HEADER_BG)
    ws.row_dimensions[3].height = 6

    # Row 4 — Stat cells
    if entity_type == "students":
        present = summary.get("student_present", summary.get("present", 0))
        late    = summary.get("student_late",    summary.get("late",    0))
    else:
        present = summary.get("staff_present",   summary.get("present", 0))
        late    = summary.get("staff_late",      summary.get("late",    0))

    est    = summary.get("estimated", 0)
    absent = max(0, est - present) if est else 0

    stats = [(f"Present: {present}", XL_SUCCESS)]
    if any_late_enabled:
        stats.append((f"Late: {late}", XL_WARNING))
    stats.append((f"Absent: {absent}", XL_ERROR))

    n_stats = len(stats)
    chunk   = max(1, num_cols // n_stats)
    for i, (text, color) in enumerate(stats):
        start = i * chunk + 1
        end   = (i + 1) * chunk if i < n_stats - 1 else num_cols
        sc    = get_column_letter(start)
        ec    = get_column_letter(end)
        if sc != ec:
            ws.merge_cells(f"{sc}4:{ec}4")
        c            = ws[f"{sc}4"]
        c.value      = text
        c.font       = _xl_font(bold=True, size=11, color=color)
        c.fill       = _xl_fill(XL_HEADER_BG)
        c.alignment  = _xl_center()
    ws.row_dimensions[4].height = 22

    # Row 5 — Expected / Rate
    half = max(1, num_cols // 2)
    sc1  = get_column_letter(1)
    ec1  = get_column_letter(half)
    sc2  = get_column_letter(half + 1)
    ec2  = get_column_letter(num_cols)
    if sc1 != ec1:
        ws.merge_cells(f"{sc1}5:{ec1}5")
    if sc2 != ec2:
        ws.merge_cells(f"{sc2}5:{ec2}5")

    rate = summary.get("rate")
    c        = ws[f"{sc1}5"]
    c.value  = f"Expected: {est if est else '—'}"
    c.font   = _xl_font(size=10, color=XL_MUTED)
    c.fill   = _xl_fill(XL_HEADER_BG)
    c.alignment = _xl_center()

    rate_color = (XL_SUCCESS if rate and rate >= 75
                  else XL_WARNING if rate and rate >= 50
                  else XL_ERROR if rate else XL_MUTED)
    c        = ws[f"{sc2}5"]
    c.value  = f"Attendance rate: {rate}%" if rate else "Attendance rate: —"
    c.font   = _xl_font(bold=True, size=10, color=rate_color)
    c.fill   = _xl_fill(XL_HEADER_BG)
    c.alignment = _xl_center()
    ws.row_dimensions[5].height = 18

    ws.merge_cells(f"A6:{last_col}6")
    ws.row_dimensions[6].height = 8
    return 7


def _xl_write_period_section(ws, period_obj, period_rows: list,
                             start_row: int, entity_type: str = "students") -> int:
    show_timeout = bool(period_obj and period_obj.timeout_enabled)
    late_enabled = period_obj.late_enabled if period_obj else True
    accent       = XL_STUDENT if entity_type == "students" else XL_STAFF

    if not late_enabled:
        period_rows = [r for r in period_rows if r["status"] != "late"]

    if entity_type == "students":
        headers = ["#", "Student ID", "Name", "Department", "Program",
                   "Code", "Year Level", "Status", "Time In"]
    else:
        headers = ["#", "Staff ID", "Name", "Department", "Role",
                   "Status", "Time In"]

    if show_timeout:
        headers.append("Time Out")

    num_cols  = len(headers)
    last_col_l = get_column_letter(num_cols)
    period_name = period_obj.name if period_obj else "Unknown Period"

    # Period heading
    ws.merge_cells(f"A{start_row}:{last_col_l}{start_row}")
    c           = ws[f"A{start_row}"]
    c.value     = f"{period_name.upper()} — {len(period_rows)} record(s)"
    c.font      = _xl_font(bold=True, size=10, color=accent)
    c.fill      = _xl_fill(XL_SECTION_BG)
    c.alignment = _xl_left()
    c.border    = _xl_bottom_border()
    ws.row_dimensions[start_row].height = 20
    current_row = start_row + 1

    # Column headers
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=current_row, column=col, value=h)
        cell.font       = _xl_font(bold=True, color=XL_HEADER_FG)
        cell.fill       = _xl_fill(accent)
        cell.alignment  = _xl_center()
        cell.border     = _xl_border()
    ws.row_dimensions[current_row].height = 18
    current_row += 1

    # Data rows
    for i, row in enumerate(period_rows, 1):
        bg = XL_ROW_ALT if i % 2 == 0 else XL_ROW_NORMAL
        sc = _XL_STATUS_COLORS.get(row["status"], XL_MUTED)

        if entity_type == "students":
            values = [i, row["entity_id"], row["name"],
                      row.get("dept_code", "—"), row.get("col3", "—"),
                      row.get("col3", "—"), row.get("col4", "—"),
                      row["status"].upper(), row["time_in"]]
        else:
            values = [i, row["entity_id"], row["name"],
                      row.get("col3", "—"), row.get("col4", "—"),
                      row["status"].upper(), row["time_in"]]

        if show_timeout:
            values.append(row["time_out"])

        status_col = 8 if entity_type == "students" else 6

        for col, val in enumerate(values, 1):
            is_status = (col == status_col)
            is_name   = (col == 3)
            cell            = ws.cell(row=current_row, column=col, value=val)
            cell.fill       = _xl_fill(bg)
            cell.border     = _xl_border()
            cell.alignment  = _xl_left() if is_name else _xl_center()
            cell.font       = _xl_font(
                bold=is_status,
                color=sc if is_status else XL_TEXT)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

    ws.row_dimensions[current_row].height = 10
    current_row += 1
    return current_row


def _build_xl_sheet(wb, sheet_title: str, session, rows: list,
                    summary: dict, period_map: dict, periods: list,
                    entity_type: str, any_late: bool,
                    academic_period: str):
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = False

    grouped: dict = OrderedDict()
    for p in periods:
        grouped[p.id] = []
    for row in rows:
        pid = row.get("period_id")
        if pid in grouped:
            grouped[pid].append(row)
        else:
            grouped.setdefault(None, []).append(row)

    num_cols    = 10
    current_row = _xl_write_header_block(
        ws, session, summary, num_cols,
        entity_type=entity_type,
        any_late_enabled=any_late,
        academic_period=academic_period)

    for period_id, period_rows in grouped.items():
        if not period_rows:
            continue
        period_obj  = period_map.get(period_id)
        current_row = _xl_write_period_section(
            ws, period_obj, period_rows, current_row,
            entity_type=entity_type)

    ws.freeze_panes = "A7"
    _xl_autosize(ws)


def export_session_xlsx(data: dict, filepath: str):
    wb       = openpyxl.Workbook()
    atype    = data.get("attendee_type", "students")
    session  = data["session"]
    summary  = data["summary"]
    pmap     = data.get("period_map", {})
    periods  = data.get("periods", [])
    any_late = data.get("any_late_enabled", True)
    ap_str   = data.get("academic_period", "—")

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if atype in ("students", "both"):
        _build_xl_sheet(wb, "Students", session,
                        data.get("student_rows", []),
                        summary, pmap, periods,
                        entity_type="students",
                        any_late=any_late,
                        academic_period=ap_str)

    if atype in ("staff", "both"):
        _build_xl_sheet(wb, "Staff", session,
                        data.get("staff_rows", []),
                        summary, pmap, periods,
                        entity_type="staff",
                        any_late=any_late,
                        academic_period=ap_str)

    wb.save(filepath)


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

_PDF_BG      = colors.HexColor("#ffffff")
_PDF_SURFACE = colors.HexColor("#f1f5f9")
_PDF_ACCENT  = colors.HexColor("#2563eb")
_PDF_STUDENT = colors.HexColor("#0369a1")
_PDF_STAFF   = colors.HexColor("#7c3aed")
_PDF_SUCCESS = colors.HexColor("#16a34a")
_PDF_WARNING = colors.HexColor("#d97706")
_PDF_ERROR   = colors.HexColor("#dc2626")
_PDF_TEXT    = colors.HexColor("#1e293b")
_PDF_MUTED   = colors.HexColor("#64748b")
_PDF_BORDER  = colors.HexColor("#cbd5e1")

_STATUS_PDF   = {"present": _PDF_SUCCESS, "late": _PDF_WARNING, "absent": _PDF_ERROR}
_STATUS_LABEL = {"present": "PRESENT",    "late": "LATE",       "absent": "ABSENT"}


def _pdf_styles() -> dict:
    return {
        "title": ParagraphStyle(
            "title", fontName="Helvetica-Bold", fontSize=20,
            textColor=_PDF_TEXT, alignment=TA_CENTER,
            spaceAfter=2, leading=24),
        "subtitle": ParagraphStyle(
            "subtitle", fontName="Helvetica", fontSize=10,
            textColor=_PDF_MUTED, alignment=TA_CENTER,
            spaceAfter=0, leading=14),
        "section_hdr": ParagraphStyle(
            "section_hdr", fontName="Helvetica-Bold", fontSize=12,
            textColor=_PDF_TEXT, alignment=TA_LEFT,
            spaceBefore=12, spaceAfter=4, leading=14),
        "meta_value": ParagraphStyle(
            "meta_value", fontName="Helvetica-Bold", fontSize=9,
            textColor=_PDF_TEXT, alignment=TA_LEFT, leading=12),
        "stat_number": ParagraphStyle(
            "stat_number", fontName="Helvetica-Bold", fontSize=22,
            textColor=_PDF_TEXT, alignment=TA_CENTER, leading=26),
        "stat_label": ParagraphStyle(
            "stat_label", fontName="Helvetica", fontSize=8,
            textColor=_PDF_MUTED, alignment=TA_CENTER, leading=10),
        "section": ParagraphStyle(
            "section", fontName="Helvetica-Bold", fontSize=9,
            textColor=_PDF_MUTED, alignment=TA_LEFT,
            spaceBefore=10, spaceAfter=3, leading=11),
        "row_name": ParagraphStyle(
            "row_name", fontName="Helvetica-Bold", fontSize=8,
            textColor=_PDF_TEXT, alignment=TA_LEFT, leading=10),
        "row_detail": ParagraphStyle(
            "row_detail", fontName="Helvetica", fontSize=7,
            textColor=_PDF_MUTED, alignment=TA_LEFT, leading=9),
        "row_center": ParagraphStyle(
            "row_center", fontName="Helvetica", fontSize=8,
            textColor=_PDF_TEXT, alignment=TA_CENTER, leading=10),
        "footer": ParagraphStyle(
            "footer", fontName="Helvetica", fontSize=7,
            textColor=_PDF_MUTED, alignment=TA_CENTER, leading=10),
    }


def _draw_page(canv: rl_canvas.Canvas, doc, session_name: str):
    w, h = A4
    canv.saveState()
    canv.setFillColor(colors.white)
    canv.rect(0, 0, w, h, fill=1, stroke=0)
    canv.setFillColor(_PDF_ACCENT)
    canv.rect(0, h - 6, w, 6, fill=1, stroke=0)
    footer_h = 10 * mm
    canv.setFillColor(_PDF_SURFACE)
    canv.rect(0, 0, w, footer_h, fill=1, stroke=0)
    canv.setStrokeColor(_PDF_BORDER)
    canv.setLineWidth(0.5)
    canv.line(0, footer_h, w, footer_h)
    canv.setFillColor(_PDF_MUTED)
    canv.setFont("Helvetica", 7)
    canv.drawString(15 * mm, 3.5 * mm,
                    f"RFID Attendance Tracker • {session_name}")
    canv.drawRightString(w - 15 * mm, 3.5 * mm,
                         f"Page {canv._pageNumber}")
    canv.restoreState()


def _stat_block(number: str, label: str,
                accent: colors.Color, s: dict) -> Table:
    cell = Table(
        [[Paragraph(number, s["stat_number"])],
         [Paragraph(label,  s["stat_label"])]],
        colWidths=[38 * mm])
    cell.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), _PDF_BG),
        ("BOX",          (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        ("LINEBEFORE",   (0, 0), (0, -1),  3,   accent),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("LEFTPADDING",  (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return cell


def _pdf_entity_section(story, rows: list, periods: list,
                        period_map: dict, summary: dict,
                        entity_type: str, any_late: bool,
                        lm: float, rm: float, s: dict):
    """Append a complete student or staff section to story."""
    w, _ = A4
    usable_w = w - lm - rm

    accent_color = _PDF_STUDENT if entity_type == "students" else _PDF_STAFF
    label        = "STUDENTS" if entity_type == "students" else "STAFF"

    if entity_type == "students":
        present = summary.get("student_present", 0)
        late    = summary.get("student_late",    0)
    else:
        present = summary.get("staff_present", 0)
        late    = summary.get("staff_late",    0)

    est    = summary.get("estimated", 0)
    absent = max(0, est - present) if est else 0

    # Section header
    story.append(Paragraph(label, ParagraphStyle(
        "entity_hdr", fontName="Helvetica-Bold", fontSize=14,
        textColor=accent_color, alignment=TA_LEFT,
        spaceBefore=8, spaceAfter=4, leading=18)))
    story.append(HRFlowable(
        width="100%", color=accent_color, thickness=1.5, spaceAfter=6))

    # Stat cards
    stat_items = [_stat_block(str(present), "PRESENT", _PDF_SUCCESS, s)]
    if any_late:
        stat_items.append(_stat_block(str(late), "LATE", _PDF_WARNING, s))
    stat_items.append(_stat_block(str(absent), "ABSENT", _PDF_ERROR, s))

    n_slots = 3
    gap     = 3 * mm
    card_w  = (usable_w - (n_slots - 1) * gap) / n_slots

    while len(stat_items) < n_slots:
        empty = Table([[Paragraph("", s["stat_label"])]], colWidths=[card_w])
        empty.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), _PDF_BG)]))
        stat_items.insert(-1, empty)

    stats_row = Table([stat_items], colWidths=[card_w] * n_slots, hAlign="LEFT")
    stats_row.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), gap / 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), gap / 2),
        ("LEFTPADDING",  (0, 0), (0,  -1), 0),
        ("RIGHTPADDING", (-1, 0), (-1, -1), 0),
    ]))
    story.append(stats_row)
    story.append(Spacer(1, 4 * mm))

    # Group rows by period
    grouped: dict = OrderedDict()
    for p in periods:
        grouped[p.id] = []
    for row in rows:
        pid = row.get("period_id")
        if pid in grouped:
            grouped[pid].append(row)
        else:
            grouped.setdefault(None, []).append(row)

    for period_id, period_rows in grouped.items():
        if not period_rows:
            continue
        period_obj   = period_map.get(period_id)
        show_timeout = bool(period_obj and period_obj.timeout_enabled)
        late_enabled = period_obj.late_enabled if period_obj else True
        period_name  = period_obj.name if period_obj else "—"

        if not late_enabled:
            period_rows = [r for r in period_rows if r["status"] != "late"]

        if entity_type == "students":
            base_cols = [8*mm, 22*mm, 0, 22*mm, 18*mm, 20*mm]
            if show_timeout:
                base_cols.append(20*mm)
            base_cols[2] = usable_w - sum(base_cols)
            labels = ["#", "Student ID", "Name / Program",
                      "Year", "Status", "Time In"]
        else:
            base_cols = [8*mm, 22*mm, 0, 28*mm, 22*mm, 20*mm]
            if show_timeout:
                base_cols.append(20*mm)
            base_cols[2] = usable_w - sum(base_cols)
            labels = ["#", "Staff ID", "Name",
                      "Department", "Role / Status", "Time In"]

        if show_timeout:
            labels.append("Time Out")

        def _hrow(labels=labels, s=s):
            return [Paragraph(
                f'<font color="#64748b">{l}</font>', s["row_center"])
                for l in labels]

        story.append(Paragraph(
            f"— {period_name.upper()} ({len(period_rows)} records)",
            s["section"]))

        table_data = [_hrow()]

        for i, row in enumerate(period_rows, 1):
            status_color = _STATUS_PDF.get(row["status"], _PDF_MUTED)
            hex_sc       = status_color.hexval()

            if entity_type == "students":
                name_cell = Table(
                    [[Paragraph(row["name"], s["row_name"])],
                     [Paragraph(
                         f'{row.get("dept_code","—")} • {row.get("col3","—")}',
                         s["row_detail"])]],
                    colWidths=[base_cols[2] - 4])
                name_cell.setStyle(TableStyle([
                    ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                    ("TOPPADDING",    (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]))
                status_pill = Table(
                    [[Paragraph(
                        f'<font color="#{hex_sc}"><b>'
                        f'{_STATUS_LABEL.get(row["status"], row["status"].upper())}'
                        f'</b></font>',
                        s["row_center"])]],
                    colWidths=[base_cols[4] - 4])
                status_pill.setStyle(TableStyle([
                    ("BACKGROUND",   (0, 0), (-1, -1), _PDF_BG),
                    ("BOX",          (0, 0), (-1, -1), 0.6, status_color),
                    ("TOPPADDING",   (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
                    ("LEFTPADDING",  (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]))
                data_row = [
                    Paragraph(str(i), s["row_center"]),
                    Paragraph(str(row["entity_id"]), s["row_center"]),
                    name_cell,
                    Paragraph(row.get("col4", "—"), s["row_center"]),
                    status_pill,
                    Paragraph(row["time_in"], s["row_center"]),
                ]
            else:
                name_cell = Table(
                    [[Paragraph(row["name"], s["row_name"])]],
                    colWidths=[base_cols[2] - 4])
                name_cell.setStyle(TableStyle([
                    ("LEFTPADDING",   (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
                    ("TOPPADDING",    (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]))
                role_status = Table(
                    [[Paragraph(
                        f'{row.get("col4","—")} · '
                        f'<font color="#{hex_sc}"><b>'
                        f'{_STATUS_LABEL.get(row["status"], row["status"].upper())}'
                        f'</b></font>',
                        s["row_center"])]],
                    colWidths=[base_cols[4] - 4])
                role_status.setStyle(TableStyle([
                    ("TOPPADDING",   (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
                    ("LEFTPADDING",  (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]))
                data_row = [
                    Paragraph(str(i), s["row_center"]),
                    Paragraph(str(row["entity_id"]), s["row_center"]),
                    name_cell,
                    Paragraph(row.get("col3", "—"), s["row_center"]),
                    role_status,
                    Paragraph(row["time_in"], s["row_center"]),
                ]

            if show_timeout:
                data_row.append(Paragraph(row["time_out"], s["row_center"]))

            table_data.append(data_row)

        t = Table(table_data, colWidths=base_cols, repeatRows=1, hAlign="LEFT")
        row_styles = [
            ("BACKGROUND",   (0, 0), (-1, 0), _PDF_SURFACE),
            ("LINEBELOW",    (0, 0), (-1, 0), 0.8, accent_color),
            ("TOPPADDING",   (0, 0), (-1, 0), 5),
            ("BOTTOMPADDING",(0, 0), (-1, 0), 5),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",    (0, 1), (-1, -1), 0.3, _PDF_BORDER),
        ]
        for i in range(1, len(table_data)):
            bg = _PDF_SURFACE if i % 2 == 0 else _PDF_BG
            row_styles.append(("BACKGROUND",    (0, i), (-1, i), bg))
            row_styles.append(("TOPPADDING",    (0, i), (-1, i), 5))
            row_styles.append(("BOTTOMPADDING", (0, i), (-1, i), 5))

        t.setStyle(TableStyle(row_styles))
        story.append(KeepTogether([t]))
        story.append(Spacer(1, 4 * mm))


def export_session_pdf(data: dict, filepath: str):
    session  = data["session"]
    summary  = data["summary"]
    pmap     = data.get("period_map", {})
    periods  = data.get("periods", [])
    atype    = data.get("attendee_type", "students")
    any_late = data.get("any_late_enabled", True)

    s        = _pdf_styles()
    w, h     = A4
    lm = rm  = 15 * mm
    tm       = 22 * mm
    bm       = 16 * mm

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=lm, rightMargin=rm,
        topMargin=tm, bottomMargin=bm)

    session_name = session.name

    def make_page(canv, doc):
        _draw_page(canv, doc, session_name)

    story = []

    # ── Cover header ──────────────────────────────────────────────────
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("ATTENDANCE REPORT", s["title"]))
    story.append(Spacer(1, 1 * mm))
    story.append(Paragraph(
        f"{session.date.strftime('%B %d, %Y')} • {session.name} • "
        f"{data.get('academic_period', '')}",
        s["subtitle"]))
    story.append(Spacer(1, 5 * mm))
    story.append(HRFlowable(
        width="100%", color=_PDF_BORDER, thickness=0.5, spaceAfter=6))

    # ── Student section ───────────────────────────────────────────────
    if atype in ("students", "both"):
        _pdf_entity_section(
            story,
            rows=data.get("student_rows", []),
            periods=periods, period_map=pmap,
            summary=summary, entity_type="students",
            any_late=any_late, lm=lm, rm=rm, s=s)

    # ── Divider between sections ──────────────────────────────────────
    if atype == "both":
        story.append(Spacer(1, 6 * mm))
        story.append(HRFlowable(
            width="100%", color=_PDF_BORDER,
            thickness=1, spaceAfter=6))

    # ── Staff section ─────────────────────────────────────────────────
    if atype in ("staff", "both"):
        _pdf_entity_section(
            story,
            rows=data.get("staff_rows", []),
            periods=periods, period_map=pmap,
            summary=summary, entity_type="staff",
            any_late=any_late, lm=lm, rm=rm, s=s)

    # ── Footer ────────────────────────────────────────────────────────
    story.append(HRFlowable(
        width="100%", color=_PDF_BORDER, thickness=0.5, spaceBefore=4))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        s["footer"]))

    doc.build(story, onFirstPage=make_page, onLaterPages=make_page)